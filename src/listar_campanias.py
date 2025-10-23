import sys
from pathlib import Path

# Configurar package para imports consistentes y PyInstaller compatibility
if __package__ in (None, ""):
    sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
    __package__ = "src"

from .excel_utils import agregar_datos, crear_o_cargar_libro_excel, obtener_o_crear_hoja, limpiar_hoja_desde_fila
from .shared.utils.legacy_utils import data_path, notify
from .shared.logging.logger import get_logger
from .utils import (
    crear_contexto_navegador,
    configurar_navegador,
    navegar_a_reportes,
    obtener_total_paginas,
    navegar_siguiente_pagina,
)
from .autentificacion import login
from .infrastructure.api import API

from playwright.sync_api import sync_playwright, Page
import re

# Rutas
ARCHIVO_BUSQUEDA = data_path("Busqueda.xlsx")

logger = get_logger()


def extraer_id_de_url(url: str) -> str:
    """
    Extrae el ID de campaña de una URL

    Ejemplo: /report/campaign/12345/ -> 12345
    """
    match = re.search(r'/campaign/(\d+)', url)
    if match:
        return match.group(1)
    return ""


def extraer_datos_campania_de_listitem(listitem_locator, page: Page) -> list[str]:
    """
    Extrae los datos de una campaña desde un listitem de la lista de informes

    Args:
        listitem_locator: Locator del listitem de la página de informes
        page: Página de Playwright

    Returns:
        Lista con los datos: ['', nombre, id, fecha, total_enviado, abierto, clics]
    """
    try:
        logger.debug("🔍 Iniciando extracción de datos de listitem")

        # Obtener el primer link que apunta a /report/campaign/ID/
        # Usando locators modernos
        campaign_link = listitem_locator.locator('a[href*="/report/campaign/"]').first

        # Verificar si existe
        if campaign_link.count() == 0:
            logger.warning("⚠️ No se encontró link de campaña en el listitem")
            return []

        # El primer link es el nombre de la campaña
        nombre = campaign_link.inner_text().strip()
        href = campaign_link.get_attribute('href') or ""
        id_campania = extraer_id_de_url(href)

        logger.debug(f"📝 Nombre extraído: {nombre}, ID: {id_campania}")

        # Obtener todo el texto del listitem
        full_text = listitem_locator.inner_text()
        logger.debug(f"📄 Texto completo del listitem: {full_text[:200]}...")

        # El texto después del nombre contiene: Tipo Fecha Listas Emails Abiertos Clics
        # Ejemplo: "20251010_Com_Novedades_SIRAJ2 Clásica 10/10/25 08:32 Equipo_Minsait , ... 8.140 2.426 0"

        # Extraer fecha (formato DD/MM/YY HH:MM o DD/MM/YY)
        fecha_match = re.search(r'(\d{2}/\d{2}/\d{2})\s*(\d{2}:\d{2})?', full_text)
        if fecha_match:
            fecha = fecha_match.group(1)
            if fecha_match.group(2):
                fecha = f"{fecha} {fecha_match.group(2)}"
        else:
            fecha = ""

        logger.debug(f"📅 Fecha extraída: {fecha}")

        # Extraer los números al final (Emails, Abiertos, Clics)
        # Estrategia: buscar el último bloque de texto que contiene solo números separados por espacios
        # Ejemplo: "... Lista1 , Lista2 8.140 2.426 0" -> queremos "8.140 2.426 0"

        # Primero, eliminar todos los links para quedarnos solo con el texto
        links = listitem_locator.locator('a').all()
        text_only = full_text
        for link in links:
            link_text = link.inner_text()
            text_only = text_only.replace(link_text, '')

        # Ahora buscar los últimos números (después de la fecha)
        # Patrón: buscar grupos de 3 números al final (pueden tener puntos como separadores)
        match = re.search(r'(\d{1,3}(?:\.\d{3})*|\d+)\s+(\d{1,3}(?:\.\d{3})*|\d+)\s+(\d{1,3}(?:\.\d{3})*|\d+)\s*$', text_only)

        if match:
            total_enviado = match.group(1).replace('.', '')
            abierto = match.group(2).replace('.', '')
            clics = match.group(3).replace('.', '')
            logger.debug(f"🔢 Números extraídos: Enviados={total_enviado}, Abiertos={abierto}, Clics={clics}")
        else:
            logger.warning(f"⚠️ No se encontró el patrón de números al final del texto")
            logger.debug(f"Texto sin links: {text_only[:200]}...")
            total_enviado = "0"
            abierto = "0"
            clics = "0"

        # Calcular "No abierto"
        try:
            no_abierto = str(int(total_enviado) - int(abierto))
        except:
            no_abierto = "0"

        logger.debug(
            f"✅ Datos extraídos exitosamente",
            extra={
                "nombre": nombre,
                "id": id_campania,
                "fecha": fecha,
                "total_enviado": total_enviado,
                "abierto": abierto,
                "no_abierto": no_abierto,
                "clics": clics
            }
        )

        return ['', nombre, id_campania, fecha, total_enviado, abierto, no_abierto]

    except Exception as e:
        logger.error(f"❌ Error extrayendo datos de campaña: {e}", extra={"error": str(e)})
        return []


def extraer_campanias_de_pagina(page: Page) -> list[list[str]]:
    """
    Extrae todas las campañas de la página actual de informes

    Args:
        page: Página de Playwright

    Returns:
        Lista de campañas con sus datos (sin duplicados por ID)
    """
    logger.info("🔍 Iniciando extracción de campañas de la página actual")
    campanias = []
    ids_vistos = set()  # Para evitar duplicados

    try:
        # Esperar a que la lista se cargue
        logger.debug("⏳ Esperando a que se cargue la lista de informes")
        page.wait_for_selector('ul li', timeout=15000)
        page.wait_for_timeout(1000)  # Espera adicional para asegurar carga completa

        logger.debug("✅ Lista de informes cargada")

        # Usar selectores modernos de Playwright
        # Estrategia: buscar todos los li que contienen un link a /report/campaign/
        # y excluir el primero que es el encabezado
        all_items = page.locator('li').filter(has=page.locator('a[href*="/report/campaign/"]'))

        # Obtener el count
        count = all_items.count()
        logger.info(f"✅ Total de elementos con links de campañas: {count}")

        # Obtener todos los elementos y filtrar manualmente los que son campañas reales
        # Estrategia mejorada: buscar elementos que tengan el patrón completo de una fila de campaña
        campaign_listitems = []
        for i in range(count):
            item = all_items.nth(i)
            text = item.inner_text()

            # Criterios para identificar una fila de campaña real:
            # 1. Debe tener una fecha en formato DD/MM/YY (obligatorio - identifica campañas reales)
            # 2. Debe tener al menos 1 número al final (pueden ser 0 0 0, o 1234, etc.)
            # 3. Debe tener longitud suficiente y no ser solo un fragmento
            tiene_fecha = bool(re.search(r'\d{2}/\d{2}/\d{2}', text))
            tiene_numeros_final = bool(re.search(r'\d+[\s,]*\d*[\s,]*\d*\s*$', text))
            longitud_suficiente = len(text.strip()) > 30

            # Verificar que NO sea un elemento anidado (no debe tener saltos de línea múltiples)
            no_es_anidado = text.count('\n') <= 3

            if tiene_fecha and tiene_numeros_final and longitud_suficiente and no_es_anidado:
                campaign_listitems.append(item)
                logger.debug(f"✅ Campaña válida encontrada en índice {i}: {text[:50]}...")
            else:
                # Logging detallado para debug - mostrar en consola los descartados
                if tiene_fecha and longitud_suficiente:  # Candidatos válidos que fueron descartados
                    print(f"⚠️ DESCARTADO [{i}]: fecha={tiene_fecha}, numeros={tiene_numeros_final}, longitud={len(text.strip())}, saltos={text.count(chr(10))}")
                    print(f"   Texto: {text[:100]}")
                logger.debug(f"⚠️ Elemento descartado en índice {i}: tiene_fecha={tiene_fecha}, tiene_numeros={tiene_numeros_final}, longitud={len(text.strip())}, saltos_linea={text.count(chr(10))}")

        logger.info(f"✅ Listitems de campañas reales encontrados: {len(campaign_listitems)}")

        for i, listitem in enumerate(campaign_listitems, 1):
            try:
                logger.debug(f"📖 Procesando campaña {i}/{len(campaign_listitems)}")
                datos = extraer_datos_campania_de_listitem(listitem, page)

                if datos and len(datos) >= 3:  # Validar que tiene datos suficientes
                    id_campania = datos[2]  # El ID está en la posición 2

                    # Verificar si ya vimos este ID (evitar duplicados)
                    if id_campania in ids_vistos:
                        logger.warning(f"⚠️ Campaña duplicada detectada (ID: {id_campania}), omitiendo...")
                        continue

                    ids_vistos.add(id_campania)
                    campanias.append(datos)
                    logger.info(f"✅ Campaña {len(campanias)} extraída: {datos[1]} (ID: {datos[2]})")
                else:
                    logger.warning(f"⚠️ No se pudieron extraer datos válidos de la campaña {i}")

            except Exception as e:
                logger.warning(f"⚠️ Error procesando listitem {i}: {e}")
                continue

        logger.success(f"✅ Extracción completada: {len(campanias)} campañas únicas extraídas de la página")

    except Exception as e:
        logger.error(f"❌ Error extrayendo campañas de la página: {e}")

    return campanias


def guardar_datos_en_excel(informe_detalle: list[list[str]], archivo_busqueda: str):
    """
    Guarda los datos en el archivo Excel, usando la primera hoja por defecto
    y ajusta automáticamente el ancho de las columnas
    """
    try:
        logger.info("🚀 Iniciando guardado de datos en Excel", extra={"archivo": archivo_busqueda, "registros": len(informe_detalle)})

        wb = crear_o_cargar_libro_excel(archivo_busqueda)
        encabezados = ["Buscar", "Nombre", "ID Campaña", "Fecha", "Total enviado", "Abierto", "No abierto"]

        # Obtener o crear la hoja "Sheet"
        ws = obtener_o_crear_hoja(wb, "Sheet")
        logger.info(f"📝 Hoja obtenida/creada: {ws.title}")

        # Limpiar hoja desde la primera fila
        logger.info("🧹 Limpiando hoja")
        limpiar_hoja_desde_fila(ws, fila_inicial=1)

        # Agregar encabezados
        logger.info("🏷️ Agregando encabezados")
        ws.append(encabezados)

        # Agregar datos
        registros_agregados = agregar_datos(ws, datos=informe_detalle)
        logger.info(f"📊 Datos agregados: {registros_agregados} registros")

        # Ajustar automáticamente el ancho de las columnas
        from openpyxl.utils import get_column_letter

        # Iterar por cada columna usando índices
        logger.info("📐 Ajustando ancho de columnas")
        for col_idx in range(1, ws.max_column + 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)

            # Revisar todas las celdas de esta columna
            for row_idx in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            # Ajustar el ancho (agregar un poco de padding)
            adjusted_width = min(max_length + 2, 50)  # Máximo 50 caracteres
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(archivo_busqueda)
        logger.success(f"✅ Archivo guardado exitosamente: {archivo_busqueda}")
        logger.info(f"📈 Se agregaron {registros_agregados} registros al archivo")

    except Exception as e:
        logger.error(f"❌ Error guardando archivo Excel: {e}")
        print(f"Error guardando archivo Excel: {e}")


def procesar_todas_las_paginas(page: Page) -> list[list[str]]:
    """
    Procesa todas las páginas de reportes y extrae todas las campañas (sin duplicados globales)

    Args:
        page: Página de Playwright

    Returns:
        Lista con todos los datos de campañas únicas
    """
    logger.info("🔍 Iniciando procesamiento de todas las páginas")
    todas_campanias = []
    ids_globales = set()  # Para evitar duplicados entre páginas

    try:
        # Obtener número total de páginas
        logger.debug("📄 Obteniendo número total de páginas")
        total_paginas = obtener_total_paginas(page)
        logger.info(f"📚 Total de páginas a procesar: {total_paginas}")

        # Procesar cada página
        for pagina_actual in range(1, total_paginas + 1):
            logger.info(f"📖 === PROCESANDO PÁGINA {pagina_actual} DE {total_paginas} ===")

            # Extraer campañas de la página actual
            campanias_pagina = extraer_campanias_de_pagina(page)

            # Filtrar duplicados globales (entre páginas)
            campanias_nuevas = 0
            for campania in campanias_pagina:
                if len(campania) >= 3:
                    id_campania = campania[2]
                    if id_campania not in ids_globales:
                        ids_globales.add(id_campania)
                        todas_campanias.append(campania)
                        campanias_nuevas += 1
                    else:
                        logger.debug(f"⚠️ Campaña duplicada entre páginas (ID: {id_campania}), omitiendo...")

            logger.info(
                f"✅ Página {pagina_actual} completada",
                extra={
                    "campanias_en_pagina": len(campanias_pagina),
                    "campanias_nuevas": campanias_nuevas,
                    "total_acumulado": len(todas_campanias)
                }
            )

            # Navegar a la siguiente página si no es la última
            if pagina_actual < total_paginas:
                logger.debug(f"➡️ Navegando a página {pagina_actual + 1}")
                exito = navegar_siguiente_pagina(page, pagina_actual)
                if not exito:
                    logger.warning(f"⚠️ No se pudo navegar a página {pagina_actual + 1}, finalizando procesamiento")
                    break

                # Espera adicional después de navegar
                page.wait_for_timeout(2000)
                logger.debug(f"✅ Navegación a página {pagina_actual + 1} completada")

        logger.success(f"🎉 Procesamiento completo: {len(todas_campanias)} campañas extraídas de {pagina_actual} páginas")

    except Exception as e:
        logger.error(f"❌ Error procesando páginas: {e}")

    return todas_campanias


def main():
    """
    Función principal del programa de listado de campañas
    Usa API para obtener IDs y scraping para completar los datos
    """
    logger.info("🚀 Iniciando programa de listado de campañas (modo híbrido: API + Scraping)")

    try:
        with sync_playwright() as p:
            # Configurar navegador
            logger.info("🌐 Configurando navegador Playwright")
            browser = configurar_navegador(p, extraccion_oculta=False)
            context = crear_contexto_navegador(browser, extraccion_oculta=False)
            page = context.new_page()
            logger.success("✅ Navegador configurado correctamente")

            # Login
            logger.info("🔐 Iniciando proceso de autenticación")
            login(page, context)
            logger.success("✅ Sesión iniciada correctamente")

            # Navegar a reportes
            logger.info("📊 Navegando a sección de reportes")
            navegar_a_reportes(page)
            logger.success("✅ Navegación a reportes completada")

            # Esperar a que la página cargue completamente
            logger.debug("⏳ Esperando carga completa de la página")
            page.wait_for_load_state("networkidle", timeout=60000)
            page.wait_for_timeout(3000)  # Espera adicional para asegurar carga completa
            logger.debug("✅ Página cargada completamente")

            # Procesar todas las páginas y extraer campañas con scraping
            logger.info("📥 Iniciando extracción de campañas mediante scraping")
            informe = procesar_todas_las_paginas(page)
            logger.info(f"📊 Total de campañas extraídas mediante scraping: {len(informe)}")

            # Guardar en Excel
            if informe:
                logger.info("💾 Guardando datos en archivo Excel")
                guardar_datos_en_excel(informe, ARCHIVO_BUSQUEDA)
                logger.success("✅ Programa completado exitosamente")
                notify("Listado de Campañas", f"Se extrajeron {len(informe)} campañas correctamente", "info")
            else:
                logger.warning("⚠️ No se encontraron campañas para guardar")
                notify("Listado de Campañas", "No se encontraron campañas", "warning")

            # Cerrar navegador
            logger.debug("🔚 Cerrando navegador")
            browser.close()
            logger.info("✅ Navegador cerrado correctamente")

    except Exception as e:
        logger.error(f"❌ Error crítico en el programa: {e}", extra={"error": str(e)})
        print(f"Error crítico en el programa: {e}")
        notify("Error", f"Error crítico: {e}", "error")
        raise


if __name__ == "__main__":
    main()
