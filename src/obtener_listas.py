from playwright.sync_api import sync_playwright, Page
from datetime import datetime
from .utils import configurar_navegador, crear_contexto_navegador, obtener_total_paginas_listas, navegar_siguiente_pagina_listas, cambiar_items_por_pagina_listas, load_config, data_path, storage_state_path, notify, get_timeouts, safe_goto, safe_wait_for_element
from .autentificacion import login
import pandas as pd
from openpyxl import Workbook, load_workbook
import os
import re

# Rutas
ARCHIVO_LISTAS = data_path("Buscar_Lista.xlsx")

def cargar_ultimo_termino_busqueda(archivo_listas: str) -> list[str]:
	"""
	Carga el último término de búsqueda desde el archivo Excel
	"""
	try:
		df = pd.read_excel(archivo_listas, engine="openpyxl")
		terminos = ["", ""]

		# Solo si hay filas y existen las columnas esperadas, extrae la última
		if not df.empty and {'Nombre', 'Suscriptores'}.issubset(df.columns):
			ultima_fila = df.iloc[-1]
			terminos = [str(ultima_fila.get('Nombre', '')).strip(),
						str(ultima_fila.get('Suscriptores', '')).strip()]
		
		return terminos
	except Exception as e:
		print(f"Error al cargar términos de búsqueda: {e}")
		return ["", ""]

def inicializar_navegacion_listas(page: Page):
	"""
	Navega a la sección de listas y espera a que cargue
	"""
	timeouts = get_timeouts()
    
	# Navegar a la página de listas
	try:
		page.goto("https://acumbamail.com/app/lists/")
		page.wait_for_load_state("domcontentloaded")
	except Exception as e:
		print(f"⚠️ No se pudo navegar a listas: {e}")

	# Esperar elemento concreto de la lista de listas
	try:
		# Usar el selector correcto encontrado durante debugging
		lista_listas = page.locator('.item')
		lista_listas.first.wait_for(timeout=timeouts['elements'])
		print(f"✅ Encontrados {lista_listas.count()} elementos de lista")
	except Exception as e:
		print(f"⚠️ Advertencia al cargar listas: {e}")
		# Esperar un poco más por si acaso
		page.wait_for_timeout(2000)

def extraer_datos_lista(elemento_lista, indice: int) -> list[str]:
	"""
	Extrae los datos de una lista específica del elemento
	"""
	try:
		# Intentar extraer información de la lista
		# Adaptarse a diferentes estructuras posibles
		nombre_txt = ""
		suscriptores = ""
		fecha_creacion = ""
		estado = ""
		
		# Buscar el nombre de la lista
		try:
			# Basado en la estructura encontrada, el nombre parece estar al inicio del texto
			texto_completo = elemento_lista.inner_text().strip()
			lineas = texto_completo.split('\n')
			nombre_txt = lineas[0].strip() if lineas else f"Lista {indice}"
		except:
			nombre_txt = f"Lista {indice}"
		
		# Buscar el número de suscriptores
		try:
			# Basado en la estructura "X suscriptores", buscar en el texto
			texto_completo = elemento_lista.inner_text()
			# Buscar patrón "número suscriptores"
			match = re.search(r'(\d+)\s+suscriptores?', texto_completo, re.IGNORECASE)
			if match:
				suscriptores = match.group(1)
			else:
				# Fallback: buscar cualquier número
				numeros = re.findall(r'\d+', texto_completo)
				suscriptores = numeros[0] if numeros else "0"
		except:
			suscriptores = "0"
		
		# Buscar fecha de creación
		try:
			# Buscar patrón "Creada el DD/MM/YYYY"
			texto_completo = elemento_lista.inner_text()
			match = re.search(r'Creada el (\d{2}/\d{2}/\d{4})', texto_completo)
			if match:
				fecha_creacion = match.group(1)
			else:
				fecha_creacion = "N/A"
		except:
			fecha_creacion = "N/A"
		
		# Buscar estado
		try:
			estado_element = elemento_lista.locator('.status, .state, .active, [data-status]').first
			estado = estado_element.inner_text().strip()
		except:
			estado = "Activo"
		
		# Verificar si es una fila de encabezados
		if (nombre_txt.upper() in ["NOMBRE", "NAME", "LISTA", "LIST"] or 
		    suscriptores.upper() in ["SUSCRIPTORES", "SUBSCRIBERS", "MEMBERS"]):
			print(f"⚠️ Saltando fila de encabezados: {nombre_txt}, {suscriptores}")
			return []  # Retornar lista vacía para indicar que se debe saltar
		
		return ['', nombre_txt, suscriptores, fecha_creacion, estado]
	except Exception as e:
		print(f"Error extrayendo datos de lista {indice}: {e}")
		return ['', '', '', '', '']

def buscar_listas_en_pagina(page: Page, terminos: list[str], numero_pagina: int) -> tuple[list[list[str]], bool]:
	"""
	Busca listas en la página actual y retorna los datos y si encontró el término buscado
	"""
	informe_detalle = []
	encontrado = False
	buscar_todo = not terminos[0] or not terminos[1]  # Si no hay términos, buscar todo
	timeouts = get_timeouts()
	
	# Si la página fue cerrada (por el usuario o por error), salir temprano
	if getattr(page, 'is_closed', None) and page.is_closed():
		return [], False

	try:
		# Buscar diferentes selectores posibles para las listas
		posibles_selectores = [
			'.item',  # Selector principal encontrado
			'.am-responsive-table-row.item',  # Selector más específico
			'.am-responsive-table-row',  # Selector fallback
			'.lists-container .list-item',
			'#lists-container .list-item',
			'.table tbody tr',
			'.list-row',
			'[data-list]',
			'.list',
			'li[class*="list"]',
			'div[class*="list"]'
		]
		
		elementos_listas = None
		for selector in posibles_selectores:
			try:
				elementos_listas = page.locator(selector)
				if elementos_listas.count() > 0:
					print(f"Encontrado selector válido: {selector}")
					break
			except:
				continue
		
		if not elementos_listas or elementos_listas.count() == 0:
			print("⚠️ No se encontraron elementos de lista")
			return [], False
		
		total = elementos_listas.count()
		print(f"Encontrados {total} elementos de lista en la página {numero_pagina}")
		
		for o in range(total):
			datos_lista = extraer_datos_lista(elementos_listas.nth(o), o)
			
			# Si la función retorna lista vacía, significa que es una fila de encabezados, saltar
			if not datos_lista:
				continue
			
			if datos_lista[1]:  # Si tiene nombre
				nombre_txt = datos_lista[1]
				suscriptores = datos_lista[2]
				
				# Si estamos buscando todo, agregar todas las listas
				if buscar_todo:
					informe_detalle = [datos_lista] + informe_detalle
				else:
					# Verificar si es la lista buscada específica
					if nombre_txt.strip() == terminos[0] and suscriptores.strip() == terminos[1]:
						encontrado = True
						break
					
					# Agregar al inicio para mantener orden cronológico inverso
					informe_detalle = [datos_lista] + informe_detalle
		
	except Exception as e:
		# Mensaje más claro si la página/contexto fue cerrada
		try:
			cerrada = (getattr(page, 'is_closed', None) and page.is_closed())
		except Exception:
			cerrada = False
		if cerrada:
			print(f"Error procesando página {numero_pagina}: la página fue cerrada")
		else:
			print(f"Error procesando página {numero_pagina}: {e}")
	
	return informe_detalle, encontrado

def guardar_datos_en_excel(informe_detalle: list[list[str]], archivo_listas: str):
	"""
	Guarda los datos en el archivo Excel
	"""
	try:
		try:
			wb = load_workbook(archivo_listas)
			ws = wb.active
		except FileNotFoundError:
			wb = Workbook()
			ws = wb.active
			if ws is None:
				ws = wb.create_sheet(title="Listas")
			ws.append(["Buscar", "Nombre", "Suscriptores", "Fecha creacion", "Estado"])

		# Agregar datos al final
		registros_agregados = 0
		for fila in informe_detalle:
			if ws is not None and any(fila) and len(fila) >= 5:  # Solo agregar filas con datos válidos
				# Verificación adicional: asegurar que no es una fila de headers
				nombre = fila[1] if len(fila) > 1 else ""
				suscriptores = fila[2] if len(fila) > 2 else ""
				
				if (nombre.upper() not in ["NOMBRE", "NAME", "LISTA", "LIST"] and 
				    suscriptores.upper() not in ["SUSCRIPTORES", "SUBSCRIBERS"]):
					ws.append(fila)
					registros_agregados += 1
				else:
					print(f"⚠️ Fila de encabezados filtrada en guardado: {nombre}, {suscriptores}")
		
		wb.save(archivo_listas)
		
	except Exception as e:
		print(f"Error guardando archivo Excel: {e}")

def procesar_busqueda_listas(page: Page, terminos: list[str]) -> list[list[str]]:
	"""
	Función principal que coordina la búsqueda de listas
	"""
	informe_detalle = []
	buscar_todo = not terminos[0] or not terminos[1]  # Si no hay términos, buscar todo
	
	# Inicializar navegación a listas
	inicializar_navegacion_listas(page)
	
	# Buscar en la página actual (las listas generalmente están en una sola página)
	print("Procesando listas...")
	datos_pagina, encontrado = buscar_listas_en_pagina(page, terminos, 1)
	
	# Mantener orden cronológico: nuevos datos al inicio
	informe_detalle = datos_pagina + informe_detalle
	listas_totales = len(datos_pagina)
	print(f"Página 1: añadidas {len(datos_pagina)} listas (total: {listas_totales})")
	
	# Intentar cambiar a 50 elementos por página para obtener más listas
	try:
		if cambiar_items_por_pagina_listas(page, 50):
			print("✅ Cambiado a 50 elementos por página")
			# Volver a buscar en la página actual con más elementos
			datos_pagina, encontrado = buscar_listas_en_pagina(page, terminos, 1)
			informe_detalle = datos_pagina
			listas_totales = len(datos_pagina)
			print(f"Página 1 (50 items): añadidas {len(datos_pagina)} listas (total: {listas_totales})")
	except Exception as e:
		print(f"⚠️ No se pudo cambiar elementos por página: {e}")

	# Verificar si hay paginación y procesarla
	try:
		paginas_totales = obtener_total_paginas_listas(page)
		if paginas_totales > 1:
			print(f"Total de páginas de listas: {paginas_totales}")
			for numero_pagina in range(2, paginas_totales + 1):
				print(f"Procesando página {numero_pagina} de {paginas_totales}...")

				# Navegar a siguiente página
				if navegar_siguiente_pagina_listas(page, numero_pagina - 1):
					datos_pagina, encontrado_pagina = buscar_listas_en_pagina(page, terminos, numero_pagina)
					informe_detalle = datos_pagina + informe_detalle
					listas_totales += len(datos_pagina)
					print(f"Página {numero_pagina}: añadidas {len(datos_pagina)} listas (acumulado: {listas_totales})")

					# Si estamos buscando una lista específica y la encontramos, parar
					if not buscar_todo and encontrado_pagina:
						encontrado = True
						print(f"Búsqueda detenida: se encontró la última lista registrada ('{terminos[0]}')")
						break
				else:
					break
	except Exception as e:
		print(f"⚠️ Error en paginación: {e}")
		# Si no hay paginación, continuar con una sola página
		pass
	
	if buscar_todo:
		print(f"Total de listas recopiladas: {listas_totales}")
	else:
		print(f"Total de listas añadidas: {listas_totales}")
	
	return informe_detalle

def main():
	"""
	Función principal del programa de listado de listas
	"""
	# Cargar config fresca y términos de búsqueda
	config = load_config()
	url = config.get("url", "")
	url_base = config.get("url_base", "")

	terminos = cargar_ultimo_termino_busqueda(ARCHIVO_LISTAS)
	
	# Si no hay términos válidos, buscar todas las listas
	if not terminos[0] or not terminos[1]:
		terminos = ["", ""]  # Términos vacíos para buscar todo
	
	try:
		with sync_playwright() as p:
			browser = configurar_navegador(p, extraccion_oculta=False)
			context = crear_contexto_navegador(browser)
			
			page = context.new_page()
			
			safe_goto(page, url_base, "domcontentloaded")
			safe_goto(page, url, "domcontentloaded")
			
			# Realizar login
			login(page)
			context.storage_state(path=storage_state_path())

			# Procesar búsqueda de listas
			informe_detalle = procesar_busqueda_listas(page, terminos)

			# Guardar resultados
			if informe_detalle:
				guardar_datos_en_excel(informe_detalle, ARCHIVO_LISTAS)
			
			browser.close()
			notify("Proceso finalizado", f"Lista de listas de suscriptores obtenida")
			
	except Exception as e:
		print(f"Error crítico en el programa: {e}")
		raise

if __name__ == "__main__":
	main()