import sys
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from typing import Optional

# Configurar package para imports consistentes y PyInstaller compatibility
if __package__ in (None, ""):
    sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
    __package__ = "src"

from .logger import get_logger

logger = get_logger()

def crear_o_cargar_libro_excel(archivo: str) -> Workbook:
    """
    Carga un libro Excel existente o crea uno nuevo si no existe.
    """
    try:
        logger.info("📂 Cargando libro Excel", archivo=archivo)
        workbook = load_workbook(archivo)
        logger.success("✅ Libro Excel cargado exitosamente", archivo=archivo)
        return workbook
    except FileNotFoundError:
        logger.info("🆕 Creando nuevo libro Excel", archivo=archivo)
        return Workbook()

def obtener_o_crear_hoja(libro: Workbook, nombre_hoja: str, encabezados: Optional[list[str]] = None) -> Worksheet:
    """
    Obtiene una hoja existente o crea una nueva con los encabezados especificados.
    """
    try:
        logger.info("📖 Obteniendo hoja existente", nombre_hoja=nombre_hoja)
        ws = libro[nombre_hoja]
        logger.success("✅ Hoja existente obtenida", nombre_hoja=nombre_hoja)
    except KeyError:
        logger.info("📝 Creando nueva hoja", nombre_hoja=nombre_hoja)
        ws = libro.create_sheet(title=nombre_hoja)
        if encabezados:
            logger.info("🏷️ Agregando encabezados a nueva hoja", nombre_hoja=nombre_hoja, encabezados=encabezados)
            agregar_encabezados(ws, encabezados)
        logger.success("✅ Nueva hoja creada", nombre_hoja=nombre_hoja)
    return ws

def agregar_encabezados(hoja: Worksheet, encabezados: list[str]) -> None:
    """
    Agrega encabezados a una hoja de Excel.
    """
    logger.info("🏷️ Agregando encabezados", cantidad=len(encabezados))
    for col_idx, encabezado in enumerate(encabezados, 1):
        hoja.cell(row=1, column=col_idx, value=encabezado)
    logger.info("✅ Encabezados agregados", encabezados=encabezados)

def limpiar_hoja_desde_fila(hoja: Worksheet, fila_inicial: int = 2) -> None:
    """
    Elimina todas las filas desde la fila especificada hasta el final.
    """
    max_row = hoja.max_row
    logger.info("🧹 Limpiando hoja", fila_inicial=fila_inicial, max_row=max_row)
    if max_row >= fila_inicial:
        filas_a_eliminar = max_row - fila_inicial + 1
        hoja.delete_rows(fila_inicial, filas_a_eliminar)
        logger.info("✅ Hoja limpiada", filas_eliminadas=filas_a_eliminar)
    else:
        logger.info("⏭️ No se requiere limpieza", fila_inicial=fila_inicial, max_row=max_row)

def agregar_datos(hoja: Worksheet, datos: list[list[str]]) -> int:
    """
    Agrega datos a una hoja y retorna el número de registros agregados.
    """
    logger.info("📊 Iniciando agregado de datos", total_filas=len(datos))
    registros_agregados = 0
    for fila_idx, fila in enumerate(datos):
        if any(fila):  # Solo agregar filas con datos
            hoja.append(fila)
            registros_agregados += 1
        else:
            logger.debug("⏭️ Fila vacía omitida", indice=fila_idx)
    
    logger.success("✅ Datos agregados exitosamente", registros_agregados=registros_agregados)
    return registros_agregados

def crear_hoja_con_datos(wb: Workbook, nombre_hoja: str, datos: list[list[str]], encabezados: list[str]) -> None:
    """
    Crea una hoja con encabezados y datos específicos.
    """
    logger.info("🆕 Creando hoja con datos", nombre_hoja=nombre_hoja, encabezados=encabezados, total_datos=len(datos))
    ws = wb.create_sheet(title=nombre_hoja)
    agregar_encabezados(ws, encabezados)
    agregar_datos(ws, datos)
    logger.success("✅ Hoja creada con datos", nombre_hoja=nombre_hoja)
