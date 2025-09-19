#!/usr/bin/env python3
"""
Versión simplificada de obtener_listas.py que omite la paginación problemática
"""

from playwright.sync_api import sync_playwright, Page
from datetime import datetime
from .utils import configurar_navegador, crear_contexto_navegador, load_config, data_path, storage_state_path, notify, get_timeouts, safe_goto
from .autentificacion import login
import pandas as pd
from openpyxl import Workbook, load_workbook
import os
import re

# Rutas
ARCHIVO_LISTAS = data_path("Buscar_Lista.xlsx")

def cargar_ultimo_termino_busqueda(archivo_listas: str) -> list[str]:
    """
    Carga el último término de búsqueda desde el archivo Excel
    """
    try:
        df = pd.read_excel(archivo_listas, engine="openpyxl")
        terminos = ["", ""]

        # Solo si hay filas y existen las columnas esperadas, extrae la última
        if not df.empty and {'Nombre', 'Suscriptores'}.issubset(df.columns):
            ultima_fila = df.iloc[-1]
            terminos = [str(ultima_fila.get('Nombre', '')).strip(),
                        str(ultima_fila.get('Suscriptores', '')).strip()]

        return terminos
    except Exception as e:
        print(f"Error al cargar términos de búsqueda: {e}")
        return ["", ""]

def extraer_datos_lista_simple(elemento_lista, indice: int) -> list[str]:
    """
    Extrae los datos de una lista específica del elemento
    """
    try:
        texto_completo = elemento_lista.inner_text().strip()
        print(f"Texto completo del elemento {indice}: {texto_completo}")

        # Separar líneas
        lineas = texto_completo.split('\n')

        # Nombre: primera línea
        nombre_txt = lineas[0].strip() if lineas else f"Lista {indice}"

        # Suscriptores: buscar patrón "X suscriptores"
        suscriptores = "0"
        match = re.search(r'(\d+)\s+suscriptores?', texto_completo, re.IGNORECASE)
        if match:
            suscriptores = match.group(1)

        # Fecha: buscar patrón "Creada el DD/MM/YYYY"
        fecha_creacion = "N/A"
        match = re.search(r'Creada el (\d{2}/\d{2}/\d{4})', texto_completo)
        if match:
            fecha_creacion = match.group(1)

        estado = "Activo"  # Asumir activo por defecto

        return ['', nombre_txt, suscriptores, fecha_creacion, estado]
    except Exception as e:
        print(f"Error extrayendo datos de lista {indice}: {e}")
        return ['', '', '', '', '']

def obtener_listas_simple(page: Page) -> list[list[str]]:
    """
    Obtiene todas las listas visibles en la página actual sin paginación
    """
    informe_detalle = []

    try:
        # Navegar a la página de listas
        print("📋 Navegando a la página de listas...")
        page.goto("https://acumbamail.com/app/lists/")
        page.wait_for_load_state("domcontentloaded", timeout=15000)

        # Esperar un poco más para que carguen los elementos
        page.wait_for_timeout(2000)

        # Buscar elementos de listas usando el selector correcto
        elementos_listas = page.locator('.item')
        count = elementos_listas.count()

        print(f"Encontrados {count} elementos de lista")

        if count == 0:
            print("⚠️ No se encontraron listas")
            return []

        # Extraer datos de cada lista
        for i in range(count):
            datos_lista = extraer_datos_lista_simple(elementos_listas.nth(i), i)

            if datos_lista and datos_lista[1]:  # Si tiene nombre válido
                informe_detalle.append(datos_lista)
                print(f"✅ Lista {i}: {datos_lista[1]} ({datos_lista[2]} suscriptores)")

        print(f"Total de listas extraídas: {len(informe_detalle)}")

    except Exception as e:
        print(f"❌ Error obteniendo listas: {e}")

    return informe_detalle

def guardar_datos_en_excel(informe_detalle: list[list[str]], archivo_listas: str):
    """
    Guarda los datos en el archivo Excel
    """
    try:
        try:
            wb = load_workbook(archivo_listas)
            ws = wb.active
        except FileNotFoundError:
            wb = Workbook()
            ws = wb.active
            if ws is None:
                ws = wb.create_sheet(title="Listas")
            ws.append(["Buscar", "Nombre", "Suscriptores", "Fecha creacion", "Estado"])

        # Agregar datos al final
        registros_agregados = 0
        for fila in informe_detalle:
            if ws is not None and any(fila) and len(fila) >= 5:
                ws.append(fila)
                registros_agregados += 1

        wb.save(archivo_listas)
        print(f"💾 Guardados {registros_agregados} registros en {archivo_listas}")

    except Exception as e:
        print(f"❌ Error guardando archivo Excel: {e}")

def main():
    """
    Función principal del programa simplificado
    """
    config = load_config()
    url = config.get("url", "")
    url_base = config.get("url_base", "")

    try:
        with sync_playwright() as p:
            browser = configurar_navegador(p, extraccion_oculta=False)
            context = crear_contexto_navegador(browser)

            page = context.new_page()

            safe_goto(page, url_base, "domcontentloaded")
            safe_goto(page, url, "domcontentloaded")

            # Realizar login
            login(page)
            context.storage_state(path=storage_state_path())

            # Obtener listas
            informe_detalle = obtener_listas_simple(page)

            # Guardar resultados
            if informe_detalle:
                guardar_datos_en_excel(informe_detalle, ARCHIVO_LISTAS)
                print(f"✅ Proceso completado. {len(informe_detalle)} listas obtenidas.")
            else:
                print("⚠️ No se encontraron listas para guardar.")

            browser.close()
            notify("Proceso finalizado", f"Lista de listas de suscriptores obtenida")

    except Exception as e:
        print(f"❌ Error crítico en el programa: {e}")
        raise

if __name__ == "__main__":
    main()