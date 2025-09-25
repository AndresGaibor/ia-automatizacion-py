from playwright.sync_api import Page
from typing import List, Optional
from datetime import datetime
import pandas as pd

from ..models.listas import (
    ListScrapingData,
    ListTableExtraction,
    ListSearchTerms,
    ListNavigationInfo,
    ListScrapingSession,
    ListScrapingResult,
    ListExtractionConfig,
    ListElementInfo
)
from ...utils import obtener_total_paginas, navegar_siguiente_pagina, data_path
from ...logger import get_logger

class ListsScraper:
    """Scraper para extraer datos de listas de suscriptores"""

    def __init__(self):
        self.logger = get_logger()
        self.archivo_busqueda = data_path("Busqueda_Listas.xlsx")

    def cargar_ultimo_termino_busqueda(self, archivo_busqueda: str) -> ListSearchTerms:
        """
        Carga el último término de búsqueda desde el archivo Excel
        """
        try:
            df = pd.read_excel(archivo_busqueda, engine="openpyxl")
            terminos = ListSearchTerms(nombre_lista="", creacion="")

            # Solo si hay filas y existen las columnas esperadas, extrae la última
            if not df.empty and {'NOMBRE LISTA', 'CREACION'}.issubset(df.columns):
                ultima_fila = df.iloc[-1]
                terminos = ListSearchTerms(
                    nombre_lista=str(ultima_fila.get('NOMBRE LISTA', '')).strip(),
                    creacion=str(ultima_fila.get('CREACION', '')).strip()
                )

            return terminos
        except Exception as e:
            self.logger.error(f"Error al cargar términos de búsqueda: {e}")
            return ListSearchTerms(nombre_lista="", creacion="")

    def inicializar_navegacion_listas(self, page: Page) -> bool:
        """
        Navega a la sección de listas y espera a que cargue
        """
        try:
            link_lista = page.get_by_role("listitem").filter(has_text="Listas").get_by_role("link")
            link_lista.click()
            page.wait_for_load_state("networkidle")
            return True
        except Exception as e:
            self.logger.error(f"Error navegando a la sección de listas: {e}")
            return False

    def extraer_datos_lista(self, li_element) -> ListElementInfo:
        """
        Extrae los datos de una lista específica del elemento Li
        """
        try:
            campos = li_element.locator('> div')

            nombre = campos.nth(0)
            nombre_txt = nombre.nth(0).inner_text()
            suscriptores = campos.nth(1).inner_text().replace(' suscriptores', '')
            fecha_creacion = campos.nth(2).inner_text().replace('Creada el ', '')

            return ListElementInfo(
                element_index=0,  # Se establecerá fuera de esta función
                nombre_texto=nombre_txt,
                suscriptores_texto=suscriptores,
                fecha_creacion_texto=fecha_creacion,
                extraction_successful=True
            )
        except Exception as e:
            self.logger.error(f"Error extrayendo datos de lista: {e}")
            return ListElementInfo(
                element_index=0,
                nombre_texto="",
                suscriptores_texto="",
                fecha_creacion_texto="",
                extraction_successful=False
            )

    def buscar_listas_en_pagina(self, page: Page, terminos: ListSearchTerms, numero_pagina: int) -> ListTableExtraction:
        """
        Busca listas en la página actual y retorna los datos y si encontró el término buscado
        """
        informe_detalle = []
        encontrado = False
        buscar_todo = terminos.search_all

        try:
            tabla_listas = page.locator('#newsletter-lists')
            page.wait_for_load_state("networkidle")

            lis = tabla_listas.locator('> li')
            cantidad_li = lis.count()

            for i in range(1, cantidad_li):
                li = lis.nth(i)
                elemento_info = self.extraer_datos_lista(li)
                elemento_info.element_index = i

                if not elemento_info.nombre_texto:  # Si no hay nombre, no la guardes
                    continue

                datos_lista = elemento_info.to_list_data()

                # Si no estamos buscando todo, mirar si coincide con los términos
                if not buscar_todo:
                    if (datos_lista.nombre_lista.strip() == terminos.nombre_lista and
                        datos_lista.creacion.strip() == terminos.creacion):
                        encontrado = True
                        break

                # Agregar al inicio para mantener orden cronológico inverso
                informe_detalle = [datos_lista] + informe_detalle

        except Exception as e:
            self.logger.error(f"Error procesando página {numero_pagina}: {e}")

        return ListTableExtraction(
            lists_found=informe_detalle,
            search_term_found=encontrado,
            page_processed=numero_pagina,
            extraction_successful=True
        )

    def obtener_listas(self, page: Page, terminos: ListSearchTerms, config: Optional[ListExtractionConfig] = None) -> ListScrapingResult:
        """
        Función principal que coordina la búsqueda de listas
        """
        if config is None:
            config = ListExtractionConfig()

        session = ListScrapingSession(
            session_id=f"lists_{datetime.now().strftime('%Y%m%d_%H%M%S')}",
            search_terms=terminos
        )

        informe_detalle = []
        buscar_todo = terminos.search_all

        # Inicializar navegación a listas
        navigation_info = ListNavigationInfo()
        if not self.inicializar_navegacion_listas(page):
            session.add_error("No se pudo navegar a la sección de listas")
            session.complete_session()
            return ListScrapingResult(
                lists_data=[],
                session_info=session,
                search_completed=False,
                target_found=False
            )

        navigation_info.lists_section_loaded = True

        try:
            paginas_totales = obtener_total_paginas(page)
            navigation_info.total_pages = paginas_totales

            self.logger.info(f"Total de páginas de listas: {paginas_totales}")

            encontrado = False
            listas_totales = 0

            # Aplicar límite de páginas si está configurado
            max_pages = min(paginas_totales, config.max_pages) if config.max_pages else paginas_totales

            # Buscar en todas las páginas
            for numero_pagina in range(1, max_pages + 1):
                navigation_info.current_page = numero_pagina
                self.logger.info(f"Procesando página {numero_pagina} de {max_pages}...")

                datos_pagina_result = self.buscar_listas_en_pagina(page, terminos, numero_pagina)
                session.pages_processed = numero_pagina

                # Mantener orden cronológico: nuevos datos al inicio
                informe_detalle = datos_pagina_result.lists_found + informe_detalle
                listas_totales += len(datos_pagina_result.lists_found)
                session.total_lists_found = listas_totales

                encontrado = datos_pagina_result.search_term_found

                # Si estamos buscando una lista específica y la encontramos, parar
                if not buscar_todo and encontrado and config.stop_on_target_found:
                    self.logger.info(f"Búsqueda detenida: se encontró la última lista registrada ('{terminos.nombre_lista}')")
                    break

                # Navegar a siguiente página si no es la última
                if numero_pagina < max_pages:
                    if not navegar_siguiente_pagina(page, numero_pagina):
                        navigation_info.navigation_successful = False
                        session.add_error(f"No se pudo navegar a la página {numero_pagina + 1}")
                        break

            if buscar_todo:
                self.logger.info(f"Total de listas recopiladas: {listas_totales}")
            else:
                self.logger.info(f"Total de listas añadidas: {listas_totales}")

            session.complete_session()

            return ListScrapingResult(
                lists_data=informe_detalle,
                session_info=session,
                search_completed=True,
                target_found=encontrado
            )

        except Exception as e:
            session.add_error(str(e))
            session.complete_session()
            self.logger.error(f"Error en obtener_listas: {e}")
            raise

    def guardar_datos_en_excel(self, informe_detalle: List[ListScrapingData], archivo_busqueda: str) -> bool:
        """
        Guarda los datos en el archivo Excel
        """
        if not informe_detalle:
            self.logger.warning("No hay datos para guardar.")
            return False

        try:
            from openpyxl import Workbook, load_workbook

            try:
                wb = load_workbook(archivo_busqueda)
                ws = wb.active
            except FileNotFoundError:
                wb = Workbook()
                ws = wb.active
                if ws is None:
                    ws = wb.create_sheet(title="Listas")
                ws.append(["Buscar", "NOMBRE LISTA", "SUSCRIPTORES", "CREACION"])

            # Agregar datos al final
            registros_agregados = 0
            for lista_data in informe_detalle:
                if ws is not None:
                    fila = [
                        lista_data.buscar,
                        lista_data.nombre_lista,
                        lista_data.suscriptores,
                        lista_data.creacion
                    ]
                    if any(fila):  # Solo agregar filas con datos
                        ws.append(fila)
                        registros_agregados += 1

            wb.save(archivo_busqueda)
            self.logger.info(f"✅ Guardados {registros_agregados} registros en {archivo_busqueda}")
            return True

        except Exception as e:
            self.logger.error(f"Error guardando archivo Excel: {e}")
            return False

    def ejecutar_scraping_completo(
        self,
        page: Page,
        terminos: Optional[ListSearchTerms] = None,
        config: Optional[ListExtractionConfig] = None
    ) -> ListScrapingResult:
        """
        Ejecuta el scraping completo de listas con todas las características
        """
        # Cargar términos de búsqueda si no se proporcionan
        if terminos is None:
            terminos = self.cargar_ultimo_termino_busqueda(self.archivo_busqueda)

        # Si no hay términos válidos, buscar todas las listas
        if not terminos.has_search_terms:
            terminos = ListSearchTerms(nombre_lista="", creacion="")

        # Ejecutar el scraping
        resultado = self.obtener_listas(page, terminos, config)

        # Guardar datos en Excel si se completó exitosamente
        if resultado.search_completed and resultado.lists_data:
            self.guardar_datos_en_excel(resultado.lists_data, self.archivo_busqueda)

        return resultado