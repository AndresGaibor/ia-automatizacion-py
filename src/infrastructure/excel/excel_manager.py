"""Excel file management and operations."""
from typing import List, Dict, Any, Optional
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from pathlib import Path

from ...core.errors import DataProcessingError
from ...shared.logging import get_logger

logger = get_logger()


class ExcelManager:
    """Manages Excel file operations with enhanced error handling."""

    @staticmethod
    def create_workbook() -> Workbook:
        """Create a new Excel workbook."""
        try:
            wb = Workbook()
            # Remove default "Sheet" if it exists
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])
            return wb
        except Exception as e:
            raise DataProcessingError(
                f"Failed to create Excel workbook: {e}"
            ) from e

    @staticmethod
    def get_or_create_worksheet(
        wb: Workbook,
        sheet_name: str,
        headers: Optional[List[str]] = None
    ) -> Worksheet:
        """Get existing worksheet or create new one with headers."""
        try:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(title=sheet_name)

                if headers:
                    for col_idx, header in enumerate(headers, start=1):
                        ws.cell(row=1, column=col_idx, value=header)

            return ws
        except Exception as e:
            raise DataProcessingError(
                f"Failed to get/create worksheet '{sheet_name}': {e}",
                context={"sheet_name": sheet_name}
            ) from e

    @staticmethod
    def add_data_to_worksheet(
        ws: Worksheet,
        data: List[List[Any]],
        start_row: int = 2
    ) -> None:
        """Add data rows to worksheet starting from specified row."""
        try:
            for row_idx, row_data in enumerate(data, start=start_row):
                for col_idx, value in enumerate(row_data, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

            logger.info(f"Added {len(data)} rows to worksheet '{ws.title}'")
        except Exception as e:
            raise DataProcessingError(
                f"Failed to add data to worksheet '{ws.title}': {e}",
                context={"sheet_name": ws.title, "data_rows": len(data)}
            ) from e

    @staticmethod
    def create_sheet_with_data(
        wb: Workbook,
        sheet_name: str,
        data: List[List[Any]],
        headers: List[str]
    ) -> Worksheet:
        """Create a new worksheet with headers and data."""
        try:
            ws = ExcelManager.get_or_create_worksheet(wb, sheet_name, headers)
            if data:
                ExcelManager.add_data_to_worksheet(ws, data)

            logger.info(f"Created sheet '{sheet_name}' with {len(data)} rows")
            return ws
        except Exception as e:
            raise DataProcessingError(
                f"Failed to create sheet '{sheet_name}' with data: {e}",
                context={"sheet_name": sheet_name, "data_rows": len(data)}
            ) from e

    @staticmethod
    def save_workbook(wb: Workbook, filename: str) -> None:
        """Save workbook to file with proper error handling."""
        try:
            # Ensure directory exists
            file_path = Path(filename)
            file_path.parent.mkdir(parents=True, exist_ok=True)

            wb.save(filename)
            logger.info(f"Excel workbook saved successfully: {filename}")
        except Exception as e:
            raise DataProcessingError(
                f"Failed to save Excel workbook: {e}",
                context={"filename": filename}
            ) from e

    @classmethod
    def create_campaign_report(
        cls,
        general_data: List[List[str]],
        detailed_data: Dict[str, List[List[str]]],
        filename: str
    ) -> str:
        """Create comprehensive campaign report Excel file."""
        try:
            logger.info("Creating campaign report Excel file", filename=filename)

            wb = cls.create_workbook()

            # General sheet
            general_headers = [
                "Nombre", "Tipo", "Fecha envio", "Listas",
                "Emails", "Abiertos", "Clics", "URL de Correo"
            ]
            cls.create_sheet_with_data(wb, "General", general_data, general_headers)

            # Detailed sheets configuration
            sheets_config = [
                ("Abiertos", detailed_data.get('opened', []), [
                    "Proyecto", "Lista", "Correo", "Fecha apertura",
                    "País apertura", "Aperturas", "Lista", "Estado", "Calidad"
                ]),
                ("No abiertos", detailed_data.get('not_opened', []), [
                    "Proyecto", "Lista", "Correo", "Lista", "Estado", "Calidad"
                ]),
                ("Clics", detailed_data.get('clicked', []), [
                    "Proyecto", "Lista", "Correo", "Fecha primer clic",
                    "País apertura", "Lista", "Estado", "Calidad"
                ]),
                ("Hard bounces", detailed_data.get('hard_bounces', []), [
                    "Proyecto", "Lista", "Correo", "Lista", "Estado", "Calidad"
                ]),
                ("Soft bounces", detailed_data.get('soft_bounces', []), [
                    "Proyecto", "Lista", "Correo", "Lista", "Estado", "Calidad"
                ])
            ]

            # Create detailed sheets
            for sheet_name, data, headers in sheets_config:
                cls.create_sheet_with_data(wb, sheet_name, data, headers)

            cls.save_workbook(wb, filename)

            logger.info(
                f"Campaign report created successfully: {filename}",
                total_sheets=len(sheets_config) + 1,
                filename=filename
            )

            return filename

        except Exception as e:
            if isinstance(e, DataProcessingError):
                raise
            raise DataProcessingError(
                f"Failed to create campaign report: {e}",
                context={"filename": filename}
            ) from e

    @staticmethod
    def validate_file_path(filepath: str) -> bool:
        """Validate if file path is writable."""
        try:
            file_path = Path(filepath)
            # Check if parent directory exists or can be created
            file_path.parent.mkdir(parents=True, exist_ok=True)
            return True
        except Exception as e:
            raise DataProcessingError(
                f"Invalid file path '{filepath}': {e}",
                context={"filepath": filepath}
            ) from e