#!/usr/bin/env python3
"""
Versión mejorada de obtener_listas.py con mejor manejo de errores y paginación
"""

from playwright.sync_api import sync_playwright, Page
from datetime import datetime
from .utils import configurar_navegador, crear_contexto_navegador, obtener_total_paginas_listas, navegar_siguiente_pagina_listas, load_config, data_path, storage_state_path, notify, get_timeouts, safe_goto
from .autentificacion import login
import pandas as pd
from openpyxl import Workbook, load_workbook
import os
import re

# Rutas
ARCHIVO_LISTAS = data_path("Buscar_Lista.xlsx")

def cargar_ultimo_termino_busqueda(archivo_listas: str) -> list[str]:
    """
    Carga el último término de búsqueda desde el archivo Excel
    """
    try:
        df = pd.read_excel(archivo_listas, engine="openpyxl")
        terminos = ["", ""]

        # Solo si hay filas y existen las columnas esperadas, extrae la última
        if not df.empty and {'Nombre', 'Suscriptores'}.issubset(df.columns):
            ultima_fila = df.iloc[-1]
            terminos = [str(ultima_fila.get('Nombre', '')).strip(),
                        str(ultima_fila.get('Suscriptores', '')).strip()]

        return terminos
    except Exception as e:
        print(f"Error al cargar términos de búsqueda: {e}")
        return ["", ""]

def inicializar_navegacion_listas(page: Page):
    """
    Navega a la sección de listas y espera a que cargue
    """
    timeouts = get_timeouts()

    # Navegar a la página de listas
    try:
        page.goto("https://acumbamail.com/app/lists/")
        page.wait_for_load_state("domcontentloaded")
        page.wait_for_timeout(2000)  # Esperar un poco más
    except Exception as e:
        print(f"⚠️ No se pudo navegar a listas: {e}")

    # Esperar elementos de lista
    try:
        lista_listas = page.locator('.item')
        lista_listas.first.wait_for(timeout=timeouts['elements'])
        print(f"✅ Encontrados {lista_listas.count()} elementos de lista")
    except Exception as e:
        print(f"⚠️ Advertencia al cargar listas: {e}")

def cambiar_items_por_pagina_seguro(page: Page, items: int = 50) -> bool:
    """
    Intenta cambiar items por página de forma segura, sin colgarse
    """
    try:
        print(f"📊 Intentando cambiar a {items} elementos por página")

        # Buscar el select dentro de .am-items-per-page
        items_per_page_container = page.locator('.am-items-per-page')

        if items_per_page_container.count() == 0:
            print("No se encontró contenedor de items per page")
            return False

        select_element = items_per_page_container.locator('select').first

        if select_element.count() == 0:
            print("No se encontró select de items per page")
            return False

        # Verificar opciones disponibles
        opciones = select_element.locator('option')
        opciones_disponibles = []
        for i in range(opciones.count()):
            texto = opciones.nth(i).inner_text().strip()
            opciones_disponibles.append(texto)

        print(f"Opciones disponibles: {opciones_disponibles}")

        # Encontrar la mejor opción (50 o la más alta)
        items_objetivo = None
        if str(items) in opciones_disponibles:
            items_objetivo = str(items)
        else:
            # Usar la opción más alta disponible
            numeros = [int(x) for x in opciones_disponibles if x.isdigit()]
            if numeros:
                items_objetivo = str(max(numeros))

        if items_objetivo:
            print(f"Seleccionando {items_objetivo} elementos por página")

            # Usar timeout más corto para evitar colgarse
            select_element.wait_for(timeout=5000)
            select_element.select_option(label=items_objetivo)

            # Esperar recarga con timeout corto
            page.wait_for_load_state("domcontentloaded", timeout=10000)
            page.wait_for_timeout(2000)

            print(f"✅ Cambiado a {items_objetivo} elementos por página")
            return True

        return False

    except Exception as e:
        print(f"⚠️ Error cambiando items por página: {e}")
        return False

def extraer_datos_lista_mejorado(elemento_lista, indice: int) -> list[str]:
    """
    Extrae los datos de una lista específica del elemento
    """
    try:
        texto_completo = elemento_lista.inner_text().strip()

        # Separar líneas
        lineas = texto_completo.split('\n')

        # Nombre: primera línea
        nombre_txt = lineas[0].strip() if lineas else f"Lista {indice}"

        # Suscriptores: buscar patrón "X suscriptores"
        suscriptores = "0"
        match = re.search(r'(\d+)\s+suscriptores?', texto_completo, re.IGNORECASE)
        if match:
            suscriptores = match.group(1)

        # Fecha: buscar patrón "Creada el DD/MM/YYYY"
        fecha_creacion = "N/A"
        match = re.search(r'Creada el (\d{2}/\d{2}/\d{4})', texto_completo)
        if match:
            fecha_creacion = match.group(1)

        estado = "Activo"  # Asumir activo por defecto

        # Verificar si es una fila de encabezados
        if (nombre_txt.upper() in ["NOMBRE", "NAME", "LISTA", "LIST"] or
            suscriptores.upper() in ["SUSCRIPTORES", "SUBSCRIBERS", "MEMBERS"]):
            print(f"⚠️ Saltando fila de encabezados: {nombre_txt}, {suscriptores}")
            return []  # Retornar lista vacía para indicar que se debe saltar

        return ['', nombre_txt, suscriptores, fecha_creacion, estado]
    except Exception as e:
        print(f"Error extrayendo datos de lista {indice}: {e}")
        return ['', '', '', '', '']

def buscar_listas_en_pagina_mejorado(page: Page, terminos: list[str], numero_pagina: int) -> tuple[list[list[str]], bool]:
    """
    Busca listas en la página actual y retorna los datos y si encontró el término buscado
    """
    informe_detalle = []
    encontrado = False
    buscar_todo = not terminos[0] or not terminos[1]  # Si no hay términos, buscar todo

    try:
        # Buscar elementos de listas usando el selector correcto
        elementos_listas = page.locator('.item')
        total = elementos_listas.count()

        if total == 0:
            print("⚠️ No se encontraron elementos de lista")
            return [], False

        print(f"Encontrados {total} elementos de lista en la página {numero_pagina}")

        for o in range(total):
            datos_lista = extraer_datos_lista_mejorado(elementos_listas.nth(o), o)

            # Si la función retorna lista vacía, significa que es una fila de encabezados, saltar
            if not datos_lista:
                continue

            if datos_lista[1]:  # Si tiene nombre
                nombre_txt = datos_lista[1]
                suscriptores = datos_lista[2]

                # Si estamos buscando todo, agregar todas las listas
                if buscar_todo:
                    informe_detalle = [datos_lista] + informe_detalle
                else:
                    # Verificar si es la lista buscada específica
                    if nombre_txt.strip() == terminos[0] and suscriptores.strip() == terminos[1]:
                        encontrado = True
                        break

                    # Agregar al inicio para mantener orden cronológico inverso
                    informe_detalle = [datos_lista] + informe_detalle

    except Exception as e:
        print(f"Error procesando página {numero_pagina}: {e}")

    return informe_detalle, encontrado

def guardar_datos_en_excel(informe_detalle: list[list[str]], archivo_listas: str):
    """
    Guarda los datos en el archivo Excel
    """
    try:
        try:
            wb = load_workbook(archivo_listas)
            ws = wb.active
        except FileNotFoundError:
            wb = Workbook()
            ws = wb.active
            if ws is None:
                ws = wb.create_sheet(title="Listas")
            ws.append(["Buscar", "Nombre", "Suscriptores", "Fecha creacion", "Estado"])

        # Agregar datos al final
        registros_agregados = 0
        for fila in informe_detalle:
            if ws is not None and any(fila) and len(fila) >= 5:
                # Verificación adicional: asegurar que no es una fila de headers
                nombre = fila[1] if len(fila) > 1 else ""
                suscriptores = fila[2] if len(fila) > 2 else ""

                if (nombre.upper() not in ["NOMBRE", "NAME", "LISTA", "LIST"] and
                    suscriptores.upper() not in ["SUSCRIPTORES", "SUBSCRIBERS"]):
                    ws.append(fila)
                    registros_agregados += 1
                else:
                    print(f"⚠️ Fila de encabezados filtrada en guardado: {nombre}, {suscriptores}")

        wb.save(archivo_listas)
        print(f"💾 Guardados {registros_agregados} registros en {archivo_listas}")

    except Exception as e:
        print(f"Error guardando archivo Excel: {e}")

def procesar_busqueda_listas_mejorado(page: Page, terminos: list[str]) -> list[list[str]]:
    """
    Función principal que coordina la búsqueda de listas con mejor manejo de errores
    """
    informe_detalle = []
    buscar_todo = not terminos[0] or not terminos[1]

    # Inicializar navegación a listas
    inicializar_navegacion_listas(page)

    # Intentar cambiar a 50 elementos por página (sin colgarse)
    if cambiar_items_por_pagina_seguro(page, 50):
        print("✅ Cambiado a 50 elementos por página")

    # Buscar en la página actual
    print("Procesando listas...")
    datos_pagina, encontrado = buscar_listas_en_pagina_mejorado(page, terminos, 1)

    # Mantener orden cronológico: nuevos datos al inicio
    informe_detalle = datos_pagina + informe_detalle
    listas_totales = len(datos_pagina)
    print(f"Página 1: añadidas {len(datos_pagina)} listas (total: {listas_totales})")

    # Verificar si hay paginación y procesarla (con manejo de errores)
    try:
        paginas_totales = obtener_total_paginas_listas(page)
        if paginas_totales > 1:
            print(f"Total de páginas de listas: {paginas_totales}")
            for numero_pagina in range(2, paginas_totales + 1):
                print(f"Procesando página {numero_pagina} de {paginas_totales}...")

                # Navegar a siguiente página con timeout
                if navegar_siguiente_pagina_listas(page, numero_pagina - 1):
                    datos_pagina, encontrado_pagina = buscar_listas_en_pagina_mejorado(page, terminos, numero_pagina)
                    informe_detalle = datos_pagina + informe_detalle
                    listas_totales += len(datos_pagina)
                    print(f"Página {numero_pagina}: añadidas {len(datos_pagina)} listas (acumulado: {listas_totales})")

                    # Si estamos buscando una lista específica y la encontramos, parar
                    if not buscar_todo and encontrado_pagina:
                        encontrado = True
                        print(f"Búsqueda detenida: se encontró la última lista registrada ('{terminos[0]}')")
                        break
                else:
                    print(f"⚠️ No se pudo navegar a página {numero_pagina}")
                    break
    except Exception as e:
        print(f"⚠️ Error en paginación: {e}")

    if buscar_todo:
        print(f"Total de listas recopiladas: {listas_totales}")
    else:
        print(f"Total de listas añadidas: {listas_totales}")

    return informe_detalle

def main():
    """
    Función principal del programa mejorado
    """
    # Cargar config fresca y términos de búsqueda
    config = load_config()
    url = config.get("url", "")
    url_base = config.get("url_base", "")

    terminos = cargar_ultimo_termino_busqueda(ARCHIVO_LISTAS)

    # Si no hay términos válidos, buscar todas las listas
    if not terminos[0] or not terminos[1]:
        terminos = ["", ""]  # Términos vacíos para buscar todo

    try:
        with sync_playwright() as p:
            browser = configurar_navegador(p, extraccion_oculta=False)
            context = crear_contexto_navegador(browser)

            page = context.new_page()

            safe_goto(page, url_base, "domcontentloaded")
            safe_goto(page, url, "domcontentloaded")

            # Realizar login
            login(page)
            context.storage_state(path=storage_state_path())

            # Procesar búsqueda de listas con mejor manejo de errores
            informe_detalle = procesar_busqueda_listas_mejorado(page, terminos)

            # Guardar resultados
            if informe_detalle:
                guardar_datos_en_excel(informe_detalle, ARCHIVO_LISTAS)

            browser.close()
            notify("Proceso finalizado", f"Lista de listas de suscriptores obtenida")

    except Exception as e:
        print(f"Error crítico en el programa: {e}")
        raise

if __name__ == "__main__":
    main()