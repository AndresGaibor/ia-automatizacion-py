from playwright.sync_api import sync_playwright, TimeoutError as PWTimeoutError, Page
from datetime import datetime

import yaml
from openpyxl import Workbook, load_workbook
import pandas as pd
import time



REAL_UA = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 15_6) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36"
)


with open('config.yaml', 'r') as file:
	config = yaml.safe_load(file)

username = config["user"]
password = config["password"]
url = config["url"]

archivo_busqueda = "./data/Busqueda.xlsx"
archivo_informes = "./data/informes"

terminos = []
df = pd.read_excel(archivo_busqueda, engine="openpyxl")

for index, row in df.iterrows():
	buscar = row['Buscar']
	if(buscar == 'x' or buscar == 'X'):
		terminos.append([row['Nombre'], row['Listas']])



def login(page: Page):
	# page.wait_for_load_state("domcontentloaded")

	if(page.url != url):
		print("no es igual")
		try:
			page.get_by_role("button", name="Aceptar todas").click();
		except PWTimeoutError:
			pass

		page.get_by_role("link", name="Entra").click()

		page.get_by_label("Correo electrónico").fill(username)
		page.get_by_label("Contraseña").fill(password)

		# <input type="submit" value="Entrar" class="g-recaptcha signup-button" id="login-button" data-sitekey="6LeOaagZAAAAADEGihAZSe2cFNNTWgxfUM5NET9Z" data-callback="onSubmit" data-action="submit">
		with page.expect_navigation(wait_until="domcontentloaded"):
			page.get_by_role("button", name="Entrar").click()
	
	try:
		page.wait_for_selector("a[href*='/reports']", timeout=10000)
		print("Login OK")
	except PWTimeoutError:
			print("Parece que no se completó el login (¿captcha?).")
			print("Resuelve el captcha manualmente en la ventana y presiona Enter aquí…")
			input()
            # Espera a que aparezca el enlace a informes
			page.wait_for_selector("a[href*='/reports']", timeout=20000)

def main():
	with sync_playwright() as p:
		browser = p.chromium.launch(
			headless=False,
			# slow_mo=120,
			args=[
                "--disable-blink-features=AutomationControlled",
                "--no-sandbox",
            ],
			) # headless=False para ver
		
		context = browser.new_context(
            viewport={"width": 1400, "height": 900},
            user_agent=REAL_UA,
            locale="es-ES",
            timezone_id="Europe/Madrid",
            ignore_https_errors=True,
			storage_state="data/storage_state.json"
        )
		
		# context.set_default_timeout(30000)
		# context.set_default_navigation_timeout(120000)
		page = context.new_page()
		page.goto(url)
		
		login(page)
		context.storage_state(path="data/storage_state.json")
		page.get_by_role('button', name="Compartir")
		# <button class="css-f4okuk-activeTabStyle" data-testid="data-testid Tab Descargar" role="tab" aria-selected="true" type="button">Descargar</button>
		page.get_by_role('button', name="Descargar")
		input = page.get_by_label('Tipo')
		# <div class="css-1mp3i2x-input-wrapper css-14t2vqa" id="report-type">
		# 	<span id="react-select-2-live-region" class="css-7pg0cj-a11yText">
		# 	</span><span aria-live="polite" aria-atomic="false" aria-relevant="additions text" role="log" class="css-7pg0cj-a11yText">
		# 	</span><div class="css-1i88p6p"><div class="css-1q0c0d5-grafana-select-value-container">
		# 	<div class="css-8nwx1l-singleValue css-0">XLSX</div><div class=" css-1eu65zc" data-value="">
		# 	<input class="" autocapitalize="none" autocomplete="off" autocorrect="off" id="react-select-2-input" spellcheck="false" tabindex="0" type="text" aria-autocomplete="list" aria-expanded="false" aria-haspopup="true" role="combobox" aria-activedescendant="" value="" style="color: inherit; background: 0px center; opacity: 0; width: 100%; grid-area: 1 / 2; font: inherit; min-width: 2px; border: 0px; margin: 0px; outline: 0px; padding: 0px;"></div></div><div class="css-hd7v9r-input-suffix"><svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 24 24" aria-hidden="true" width="16" height="16" class="css-1d3xu67-Icon" aria-label="angle-down"><path d="M17,9.17a1,1,0,0,0-1.41,0L12,12.71,8.46,9.17a1,1,0,0,0-1.41,0,1,1,0,0,0,0,1.42l4.24,4.24a1,1,0,0,0,1.42,0L17,10.59A1,1,0,0,0,17,9.17Z"></path></svg></div></div></div>
		texto = input.locator('> div > div > div')
		texto.fill('XLSX')
		
		# page.get_by_role("link", name="Entra2").click()

		# Selector para el botón "Descargar XLSX"
		descargar_xlsx_btn = page.locator('button#getDocument')
		descargar_xlsx_btn.click()


		informe_detalle: list[list[str]] = []
		informe_detallado: list[list[str]] = []
		# <input placeholder="Buscar informe" type="text" 
		# name="search_word" class="height-32 bordered 
		# bg-color-white-1 border-radius-03 padding-left-32 padding-right-32 font-size-13 font-color-darkblue-1 line-height-20" value="">
		buscador = page.get_by_placeholder("Buscar informe");
		for termino in terminos:
			page.click("a[href*='/reports']")

			buscador.fill("");
			buscador.fill(termino[0])
			page.keyboard.press("Enter")

			page.wait_for_load_state('networkidle')

			tabla_reporte = page.locator('#newsletter-reports')

			# tabla_reporte.locator("a").first.wait_for(timeout=10000)

			tds = tabla_reporte.locator('> li')

			count6 = tds.count()

			divs = tds.first.locator('> div')

			for g in range(0, count6):
				td = tds.nth(g)
				div = td.locator('> div')
				# print(div.first.inner_html())
				nombre_txt = div.first.inner_text()
				listas = div.nth(3).inner_text()
				# print(f"{nombre_txt} {listas}")

				if(nombre_txt.strip() == termino[0] and listas.strip() == termino[1]):
					divs = td.locator('> div')

			# nombre = divs
			nombre_txt = divs.first.inner_text()
			tipo = divs.nth(1).inner_text()
			fecha_envio = divs.nth(2).inner_text()
			listas = divs.nth(3).inner_text()
			emails = divs.nth(4).inner_text()
			abiertos = divs.nth(5).inner_text()
			clics = divs.nth(6).inner_text()

			informe_detalle.append([
				nombre_txt, tipo, fecha_envio, listas, emails, abiertos, clics
			])

			divs.locator('a').first.click()
			# page.wait_for_load_state("networkidle")
			page.get_by_role('link', name="Detalles suscriptores").click()

			tabla_suscriptores = page.locator('ul').filter(has=page.locator("li", has_text="Correo electrónico"))

			suscriptores = tabla_suscriptores.locator('> li')

			count = suscriptores.count()

			for i in range(1, count):  # empieza en 1 → segundo elemento
				datos_suscriptor = suscriptores.nth(i).locator('> div')

				correo = datos_suscriptor.nth(0).inner_text()
				fecha_apertura = datos_suscriptor.nth(1).inner_text()
				pais_apertura = datos_suscriptor.nth(2).inner_text()
				aperturas = datos_suscriptor.nth(3).inner_text()
				lista = datos_suscriptor.nth(4).inner_text()
				estado = datos_suscriptor.nth(5).inner_text()
				calidad = datos_suscriptor.nth(6).inner_text()

				informe_detallado.append([
					nombre_txt, correo, fecha_apertura, pais_apertura, aperturas, lista, estado, calidad
				])

			# <a class="font-size-13 transition-02 line-height-18 display-block font-color-grey-3 font-weight-400" href="/report/campaign/3331731/url/" onclick="trackAmplitudeEvent(&quot;campaign report item clicked&quot;, {&quot;item&quot;: &quot;url tracking&quot;})">Seguimiento url's</a>
			page.get_by_role('link', name="Seguimiento url's").click()
			page.wait_for_load_state('networkidle')

			# Click en el botón de los tres puntos (ícono)
			# page.locator('div.dropleft >> div.dropdown-toggle').click()
			fil = page.locator('ul').filter(has=page.locator("span", has_text="Han hecho clic"))
			# print(fil.inner_html())
			prueb = fil.locator('> li')
			count4 = prueb.count()

			for z in range(1, count4):
				fil = page.locator('ul').filter(has=page.locator("span", has_text="Han hecho clic"))
				prueb = fil.locator('> li')
				# print(prueb.nth(z).inner_html())
				print(f"Hay {prueb.count()} filas")
				divs = prueb.nth(z).locator('> div')
				# print(divs.last.inner_html())
				prueb2 = divs.last.locator('a').first
				prueb2.click()
				prueb.nth(z).get_by_role('link', name="Detalles").click()
				
				tabla_clics = page.locator('ul').filter(has=page.locator("span", has_text="Suscriptores que han hecho clic"))
				clics = tabla_clics.locator('> li')

				count3 = clics.count()

				for k in range(1, count3):  # empieza en 1 → segundo elemento
					email = clics.nth(k).inner_text()
					for detalle in informe_detallado:
						# print(f"Correo es {email} {detalle[1]}")
						if detalle[1].strip() == email.strip():
							# Remove existing "si" if present to avoid duplicates
							if len(detalle) == 9:
								detalle.pop()
							detalle.append("SI")
						elif len(detalle) < 9:
							# Add empty string for non-matching emails
							detalle.append("")

				page.go_back(wait_until='networkidle')

			# for suscriptor in range(s):
			print("Haciendo clic para volver a la paginas de reportes")
			page.get_by_role('link', name='Emails').click()
			# page.get_by_role('link', name='Informes')
			# page.click("a[href*='/reports']")


		# Crear un libro y una hoja
		wb = Workbook()
		ws = wb.active
		if ws is not None:
			ws.title = "Hoja1"
		else:
			ws = wb.create_sheet(title="Hoja1")
		


		# Escribir datos
		ws['A1'] = "Nombre"
		ws['B1'] = "Tipo"
		ws['C1'] = "Fecha envio"
		ws['D1'] = "Listas"
		ws['E1'] = "Emails"
		ws['F1'] = "Abiertos"
		ws['G1'] = "Clics"

		for fila in informe_detalle:
			ws.append(fila)
		
		ws = wb.create_sheet(title="Hoja2")

		ws['A1'] = "Proyecto"
		ws['B1'] = "Correo"
		ws['C1'] = "Fecha apertura"
		ws['D1'] = "Pais apertura"
		ws['E1'] = "Aperturas"
		ws['F1'] = "Lista"
		ws['G1'] = "Estado"
		ws['H1'] = "Calidad"
		ws['I1'] = "Seguimiento url"

		for fila in informe_detallado:
			print(fila)
			ws.append(fila)
			
		# ws.append(["Luis", 30])
		ahora = datetime.now()
		fecha_texto = ahora.strftime("%Y%m%d%H%M")

		# Guardar archivo
		wb.save(f"{archivo_informes}_{fecha_texto}.xlsx")
		print(f"Termine, revisar archivo informes_{fecha_texto}.xlsx")
		browser.close()
if __name__ == "__main__":
	main()